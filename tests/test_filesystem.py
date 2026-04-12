"""
Unit tests for subconscious.tools.filesystem

All tests operate inside a temporary directory (inside the user home tree)
so the sandbox check passes and no real files are touched.
"""

import pathlib
import pytest
import pytest_asyncio

from subconscious.tools.filesystem import (
  read_file,
  read_range,
  search_in_file,
  search_files,
  list_directory,
  create_file,
  get_file_info,
  move_to_trash,
)


# ---------------------------------------------------------------------------
# Fixture: temporary directory inside home so _safe_path allows it
# ---------------------------------------------------------------------------

@pytest_asyncio.fixture
async def tmp_home_dir(tmp_path, monkeypatch):
  """
  Redirect pathlib.Path.home() to tmp_path so the sandbox always allows
  access during tests.
  """
  monkeypatch.setattr(pathlib.Path, "home", staticmethod(lambda: tmp_path))
  return tmp_path


# ---------------------------------------------------------------------------
# read_file
# ---------------------------------------------------------------------------

async def test_read_file_success(ctx, tmp_home_dir):
  f = tmp_home_dir / "hello.txt"
  f.write_text("Hello, world!", encoding="utf-8")
  result = await read_file(ctx, str(f))
  assert result == "Hello, world!"


async def test_read_file_not_found(ctx, tmp_home_dir):
  result = await read_file(ctx, str(tmp_home_dir / "missing.txt"))
  assert "not found" in result.lower()


async def test_read_file_is_dir(ctx, tmp_home_dir):
  result = await read_file(ctx, str(tmp_home_dir))
  assert "not a file" in result.lower()


async def test_read_file_outside_home_blocked(ctx, tmp_path, monkeypatch):
  # Point home to a *different* tmp dir so the real tmp_path is outside home
  other = tmp_path / "fake_home"
  other.mkdir()
  monkeypatch.setattr(pathlib.Path, "home", staticmethod(lambda: other))
  result = await read_file(ctx, str(tmp_path / "escape.txt"))
  assert "Access denied" in result or "outside" in result


async def test_read_file_skeleton_mode(ctx, tmp_home_dir):
  f = tmp_home_dir / "big.txt"
  # Write ~3 MB of text (3 million chars) to trigger skeleton mode
  f.write_text("A" * 3_000_000, encoding="utf-8")
  result = await read_file(ctx, str(f))
  assert "[SKELETON MODE" in result
  assert "Use read_range()" in result


# ---------------------------------------------------------------------------
# Document Readers (.docx, .xlsx, .pdf)
# ---------------------------------------------------------------------------

async def test_read_docx_mock(ctx, tmp_home_dir):
  """
  Verify docx reading. We'll create a simple docx using the library.
  """
  import docx
  f = tmp_home_dir / "test.docx"
  doc = docx.Document()
  doc.add_paragraph("Hello from Word")
  doc.save(f)
  
  result = await read_file(ctx, str(f))
  assert "Hello from Word" in result


async def test_read_xlsx_mock(ctx, tmp_home_dir):
  """
  Verify Excel reading.
  """
  import openpyxl
  f = tmp_home_dir / "test.xlsx"
  wb = openpyxl.Workbook()
  ws = wb.active
  ws.title = "Data"
  ws["A1"] = "Header"
  ws["A2"] = "Value123"
  wb.save(f)
  
  result = await read_file(ctx, str(f))
  assert "Sheet: Data" in result
  assert "Header" in result
  assert "Value123" in result


async def test_read_pdf_mock(ctx, tmp_home_dir):
  """
  PDFs are hard to generate from scratch without more dependencies (like reportlab).
  However, we can verify that the code attempts to read it and handles errors 
  or empty states gracefully if we give it a dummy.
  For a robust test, we'll check that a non-PDF file renamed to .pdf 
  triggers the pypdf error handler rather than the text decoder.
  """
  f = tmp_home_dir / "fake.pdf"
  f.write_text("not a real pdf content")
  
  result = await read_file(ctx, str(f))
  # Should trigger the internal try/except in _read_pdf
  assert "Error reading PDF" in result


# ---------------------------------------------------------------------------
# list_directory
# ---------------------------------------------------------------------------

async def test_list_directory_shows_files(ctx, tmp_home_dir):
  (tmp_home_dir / "alpha.txt").write_text("a")
  (tmp_home_dir / "beta.txt").write_text("b")
  result = await list_directory(ctx, str(tmp_home_dir))
  names = [e["name"] for e in result]
  assert "alpha.txt" in names
  assert "beta.txt" in names


async def test_list_directory_hides_dotfiles_by_default(ctx, tmp_home_dir):
  (tmp_home_dir / ".hidden").write_text("secret")
  (tmp_home_dir / "visible.txt").write_text("ok")
  result = await list_directory(ctx, str(tmp_home_dir))
  names = [e["name"] for e in result]
  assert ".hidden" not in names
  assert "visible.txt" in names


async def test_list_directory_shows_dotfiles_when_asked(ctx, tmp_home_dir):
  (tmp_home_dir / ".hidden").write_text("secret")
  result = await list_directory(ctx, str(tmp_home_dir), show_hidden=True)
  names = [e["name"] for e in result]
  assert ".hidden" in names


async def test_list_directory_not_found(ctx, tmp_home_dir):
  result = await list_directory(ctx, str(tmp_home_dir / "no_such_dir"))
  assert any("error" in e for e in result) or any("not found" in str(e).lower() for e in result)


async def test_list_directory_entries_have_type(ctx, tmp_home_dir):
  sub = tmp_home_dir / "subdir"
  sub.mkdir()
  (tmp_home_dir / "file.txt").write_text("x")
  result = await list_directory(ctx, str(tmp_home_dir))
  by_name = {e["name"]: e for e in result}
  assert by_name["subdir"]["type"] == "dir"
  assert by_name["file.txt"]["type"] == "file"


# ---------------------------------------------------------------------------
# create_file
# ---------------------------------------------------------------------------

async def test_create_file_creates_new(ctx, tmp_home_dir):
  path = str(tmp_home_dir / "new_file.txt")
  result = await create_file(ctx, path, content="created!")
  assert "created" in result.lower() or "ok" in result.lower()
  assert pathlib.Path(path).read_text() == "created!"


async def test_create_file_no_overwrite_by_default(ctx, tmp_home_dir):
  path = str(tmp_home_dir / "existing.txt")
  pathlib.Path(path).write_text("original")
  result = await create_file(ctx, path, content="new")
  assert "already exists" in result.lower() or "exist" in result.lower()
  # File content unchanged
  assert pathlib.Path(path).read_text() == "original"


async def test_create_file_overwrite_flag(ctx, tmp_home_dir):
  path = str(tmp_home_dir / "overwrite_me.txt")
  pathlib.Path(path).write_text("old")
  await create_file(ctx, path, content="new", overwrite=True)
  assert pathlib.Path(path).read_text() == "new"


async def test_create_file_creates_parents(ctx, tmp_home_dir):
  path = str(tmp_home_dir / "deep" / "nested" / "file.txt")
  result = await create_file(ctx, path, content="deep")
  assert pathlib.Path(path).exists()


# ---------------------------------------------------------------------------
# get_file_info
# ---------------------------------------------------------------------------

async def test_get_file_info_returns_dict(ctx, tmp_home_dir):
  f = tmp_home_dir / "info.txt"
  f.write_text("some content")
  result = await get_file_info(ctx, str(f))
  assert isinstance(result, dict)
  assert result.get("name") == "info.txt"
  assert result.get("size_bytes") == len("some content")


async def test_get_file_info_not_found(ctx, tmp_home_dir):
  result = await get_file_info(ctx, str(tmp_home_dir / "ghost.txt"))
  # Should return a dict with an error key or a message
  assert "error" in result or "not found" in str(result).lower()


# ---------------------------------------------------------------------------
# read_range
# ---------------------------------------------------------------------------

async def test_read_range_success(ctx, tmp_home_dir):
  f = tmp_home_dir / "lines.txt"
  content = "\n".join(f"Line {i}" for i in range(1, 101))
  f.write_text(content, encoding="utf-8")
  result = await read_range(ctx, str(f), 5, 10)
  assert "[Lines 5–10" in result
  assert "Line 5" in result
  assert "Line 10" in result


async def test_read_range_out_of_bounds(ctx, tmp_home_dir):
  f = tmp_home_dir / "short.txt"
  f.write_text("Line 1\nLine 2", encoding="utf-8")
  result = await read_range(ctx, str(f), 1, 10)
  assert "[Lines 1–2" in result  # Should cap at total lines


async def test_read_range_not_found(ctx, tmp_home_dir):
  result = await read_range(ctx, str(tmp_home_dir / "missing.txt"), 1, 5)
  assert "not found" in result.lower()


# ---------------------------------------------------------------------------
# search_in_file
# ---------------------------------------------------------------------------

async def test_search_in_file_success(ctx, tmp_home_dir):
  f = tmp_home_dir / "search.txt"
  content = "This is a test file.\nIt has multiple lines.\nTest is here.\nEnd."
  f.write_text(content, encoding="utf-8")
  result = await search_in_file(ctx, str(f), "test")
  assert "match(es) for 'test'" in result
  assert "1:" in result  # Line 1 has "test"


async def test_search_in_file_no_matches(ctx, tmp_home_dir):
  f = tmp_home_dir / "no_match.txt"
  f.write_text("No matching content here.", encoding="utf-8")
  result = await search_in_file(ctx, str(f), "xyz")
  assert "No matches found" in result


async def test_search_in_file_regex(ctx, tmp_home_dir):
  f = tmp_home_dir / "regex.txt"
  f.write_text("Line 1\nLine 2\nLine 10", encoding="utf-8")
  result = await search_in_file(ctx, str(f), r"Line \d+")
  assert "match(es)" in result
  assert "1:" in result


# ---------------------------------------------------------------------------
# search_files
# ---------------------------------------------------------------------------

async def test_search_files_by_name(ctx, tmp_home_dir):
  (tmp_home_dir / "file1.txt").write_text("content")
  (tmp_home_dir / "file2.py").write_text("code")
  result = await search_files(ctx, str(tmp_home_dir), "*.txt")
  assert len(result) == 1
  assert result[0]["name"] == "file1.txt"


async def test_search_files_with_content(ctx, tmp_home_dir):
  (tmp_home_dir / "doc1.txt").write_text("This has search term")
  (tmp_home_dir / "doc2.txt").write_text("No term here")
  result = await search_files(ctx, str(tmp_home_dir), "*.txt", "search term")
  assert len(result) == 1
  assert result[0]["name"] == "doc1.txt"
  assert "first_match_line" in result[0]


async def test_search_files_no_results(ctx, tmp_home_dir):
  result = await search_files(ctx, str(tmp_home_dir), "*.xyz")
  assert len(result) == 1
  assert "No files found" in result[0]["message"]


# ---------------------------------------------------------------------------
# move_to_trash
# ---------------------------------------------------------------------------

async def test_move_to_trash_success(ctx, tmp_home_dir):
  f = tmp_home_dir / "to_trash.txt"
  f.write_text("content")
  result = await move_to_trash(ctx, str(f))
  assert "Moved to trash" in result
  assert not f.exists()


async def test_move_to_trash_not_found(ctx, tmp_home_dir):
  result = await move_to_trash(ctx, str(tmp_home_dir / "missing.txt"))
  assert "not found" in result.lower()
